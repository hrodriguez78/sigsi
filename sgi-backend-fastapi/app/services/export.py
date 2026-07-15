import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _apply_header(ws, headers: list[str]):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _apply_row(ws, row_idx: int, values: list[Any]):
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = CELL_ALIGNMENT
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _title_sheet(ws, title: str):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, size=14, color="1565C0")
    cell.alignment = Alignment(horizontal="left")
    ws.cell(row=2, column=1, value=f"Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws.cell(row=2, column=1).font = Font(italic=True, size=9, color="888888")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20


def export_risks(data: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Riesgos"

    _title_sheet(ws, "Matriz de Riesgos")

    headers = ["Código", "Nombre", "Categoría", "Descripción", "Probabilidad",
               "Impacto", "Nivel", "Score", "Tratamiento", "Estado", "Propietario",
               "Etiquetas", "Fecha Creación"]
    _apply_header(ws, headers)

    for idx, r in enumerate(data, 3):
        _apply_row(ws, idx, [
            r.get("code", ""),
            r.get("name", ""),
            r.get("category", ""),
            r.get("description", ""),
            r.get("probability", ""),
            r.get("impact", ""),
            r.get("risk_level", ""),
            r.get("risk_score", ""),
            r.get("treatment", "") or "",
            r.get("status", ""),
            r.get("owner_id", "") or "",
            ", ".join(r.get("tags", [])),
            r.get("created_at", ""),
        ])

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_incidents(data: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Incidentes"

    _title_sheet(ws, "Registro de Incidentes")

    headers = ["Título", "Tipo", "Severidad", "Prioridad", "Estado",
               "Método Detección", "Asignado a", "Reportado por",
               "Causa Raíz", "Acciones Contención", "Lecciones",
               "Fecha Creación", "Fecha Resolución"]
    _apply_header(ws, headers)

    for idx, i in enumerate(data, 3):
        _apply_row(ws, idx, [
            i.get("title", ""),
            i.get("incident_type", ""),
            i.get("severity", ""),
            i.get("priority", ""),
            i.get("status", ""),
            i.get("detection_method", ""),
            i.get("assigned_to", "") or "",
            i.get("reported_by", "") or "",
            i.get("root_cause", "") or "",
            i.get("containment_actions", "") or "",
            i.get("lessons_learned", "") or "",
            i.get("created_at", ""),
            i.get("resolved_at", "") or "",
        ])

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_controls(data: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Controles"

    _title_sheet(ws, "Controles ISO 27001:2022")

    headers = ["Código", "Nombre", "Categoría", "Descripción", "Estado",
               "Nivel Cumplimiento", "Implementador", "Fecha Implementación",
               "Evidencias", "Fecha Creación"]
    _apply_header(ws, headers)

    for idx, c in enumerate(data, 3):
        _apply_row(ws, idx, [
            c.get("code", ""),
            c.get("name", ""),
            c.get("category", ""),
            c.get("description", ""),
            c.get("status", ""),
            c.get("compliance_level", ""),
            c.get("implementer_id", "") or "",
            c.get("implementation_date", "") or "",
            c.get("evidence_count", 0),
            c.get("created_at", ""),
        ])

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_audits(data: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Auditorías"

    _title_sheet(ws, "Registro de Auditorías")

    headers = ["Título", "Tipo", "Alcance", "Criterio", "Estado",
               "Auditor", "Email Auditor", "Fecha Planificada",
               "Fecha Inicio", "Fecha Fin", "Miembros Equipo",
                "Fecha Creación"]
    _apply_header(ws, headers)

    for idx, a in enumerate(data, 3):
        _apply_row(ws, idx, [
            a.get("title", ""),
            a.get("audit_type", ""),
            a.get("scope", ""),
            a.get("criteria", ""),
            a.get("status", ""),
            a.get("auditor_name", "") or "",
            a.get("auditor_email", "") or "",
            a.get("planned_date", "") or "",
            a.get("start_date", "") or "",
            a.get("end_date", "") or "",
            ", ".join(a.get("team_members", [])),
            a.get("created_at", ""),
        ])

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_training(data: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Capacitación"

    _title_sheet(ws, "Cursos de Capacitación")

    headers = ["Código", "Título", "Categoría", "Estado", "Duración (hrs)",
               "Instructor", "Máx. Participantes", "Inscritos",
               "Completados", "Obligatorio", "Fecha Creación"]
    _apply_header(ws, headers)

    for idx, t in enumerate(data, 3):
        _apply_row(ws, idx, [
            t.get("code", ""),
            t.get("title", ""),
            t.get("category", ""),
            t.get("status", ""),
            t.get("duration_hours", ""),
            t.get("instructor", "") or "",
            t.get("max_participants", "") or "",
            t.get("enrollments_count", 0),
            t.get("completed_count", 0),
            "Sí" if t.get("is_mandatory") else "No",
            t.get("created_at", ""),
        ])

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_documents(data: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Documentos"

    _title_sheet(ws, "Gestión Documental")

    headers = ["Código", "Título", "Tipo", "Estado", "Proceso",
               "Versión Actual", "Responsable", "Fecha Publicación",
               "Fecha Creación"]
    _apply_header(ws, headers)

    for idx, d in enumerate(data, 3):
        _apply_row(ws, idx, [
            d.get("code", ""),
            d.get("title", ""),
            d.get("document_type", ""),
            d.get("status", ""),
            d.get("process_id", "") or "",
            d.get("current_version", ""),
            d.get("owner_id", "") or "",
            d.get("published_at", "") or "",
            d.get("created_at", ""),
        ])

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_assets(data: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Activos"

    _title_sheet(ws, "Inventario de Activos")

    headers = ["Código", "Nombre", "Tipo", "Descripción", "Criticidad",
               "Estado", "Proceso", "C", "I", "D",
               "Propietario", "Fecha Creación"]
    _apply_header(ws, headers)

    for idx, a in enumerate(data, 3):
        cia = a.get("cia", {})
        _apply_row(ws, idx, [
            a.get("code", ""),
            a.get("name", ""),
            a.get("asset_type", ""),
            a.get("description", ""),
            a.get("criticality", ""),
            a.get("status", ""),
            a.get("process_id", "") or "",
            cia.get("confidentiality", "") if isinstance(cia, dict) else "",
            cia.get("integrity", "") if isinstance(cia, dict) else "",
            cia.get("availability", "") if isinstance(cia, dict) else "",
            a.get("owner_id", "") or "",
            a.get("created_at", ""),
        ])

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_raci(data: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Matriz RACI"

    _title_sheet(ws, "Matriz RACI - Responsabilidades")

    headers = ["Nombre", "Descripción", "Procesos", "Roles", "Fecha Creación"]
    _apply_header(ws, headers)

    for idx, m in enumerate(data, 3):
        _apply_row(ws, idx, [
            m.get("name", ""),
            m.get("description", ""),
            str(len(m.get("process_ids", []))),
            ", ".join(m.get("role_names", [])),
            m.get("created_at", ""),
        ])

    _auto_width(ws)

    for m in data:
        assignments = m.get("assignments", {})
        if assignments:
            ws2 = wb.create_sheet(title=m.get("name", "Detalle")[:31])
            roles = m.get("role_names", [])
            procs = m.get("process_ids", [])
            ws2.cell(row=1, column=1, value="Proceso \\ Rol")
            for ci, role in enumerate(roles, 2):
                ws2.cell(row=1, column=ci, value=role)
            for ri, pid in enumerate(procs, 2):
                ws2.cell(row=ri, column=1, value=pid)
                for ci, role in enumerate(roles, 2):
                    ws2.cell(row=ri, column=ci, value=assignments.get(pid, {}).get(role, ""))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


EXPORTERS = {
    "risks": export_risks,
    "incidents": export_incidents,
    "controls": export_controls,
    "audits": export_audits,
    "training": export_training,
    "documents": export_documents,
    "assets": export_assets,
    "raci": export_raci,
}
